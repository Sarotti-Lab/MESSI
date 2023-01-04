# -*- coding: utf-8 -*-
"""
Created on Nov 04 2022
@author: María M. Zanardi & Ariel M. Sarotti

-----------------------------------------------------------------------------------
-----------------------------------------------------------------------------------
          This program lets you to perform MESSI calculation automatically.
The program allows the use of H and C data. Although it can give partial results
using some subset of data, it is recommended to use the complete data set.

*Conditions for its use:

        You must place in a folder: all the outputs containing NMR and Single Point
calculations, corresponding to all posible conformations and isomers under
analysis (in format *.log or *.out). The names of the file will necessarily
be coded as IsomerNumber_*.

        Additionally the folder must also contain an excel file consigning the
experimental data an labels ordered as follows:
    -experimental chemical shifts and labels of the atoms (sheet namme = 'shifts')
-------------------------------------------------------------------------------------
-------------------------------------------------------------------------------------
"""

#__version__ = "0.1.9"
#__author__ = 'María M. Zanardi & Ariel M. Sarotti'

import glob
import os
import scipy.stats as stats
import copy
import pandas as pd
import numpy as np
import tkinter as tk
import webbrowser
import shutil

from pathlib import Path
from sys import exit
from tkinter import Tk, filedialog
from math import isnan
from sklearn import linear_model
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from pathlib import Path



def GUI():
    global root
    root = tk.Tk()      #define the widget
    root.wm_title("MESSI")
    root.geometry("750x230")

    title = tk.Label( root, text= 'MESSI')
    title.place(x=20, y= 5)
    title.config(font = ("Times", "35", "bold"))

    title = tk.Label( root, text= 'Multi Ensamble Strategy for Structural Identification')
    title.place(x=190, y= 15)
    title.config(font = ("Times", "11"),justify='right')

    title = tk.Label( root, text= '**************************************')
    title.place(x=190, y= 34)
    title.config(font = ("Times", "11", 'bold'),justify='right')
    #-----------------------------------------------------------------
    button_userguide = tk.Button(root, text='User Guide',
                               command=user_guide).place(x=650,y=20)
    #-----------------------------------------------------------------
    button_example = tk.Button(root, text='Create Example',
                               command=example).place(x=635,y=55)
    #-----------------------------------------------------------------
    button_end = tk.Button(root, text='Exit',
                           command=exit).place(x=20,y=190)
    #-----------------------------------------------------------------
    button_xlsx = tk.Button(root, text='Select Excel',
                           font=("Helvetica", 10),
                           command=lambda: select_xlsx ([button_xlsx,button_dir]),
                           state='disabled')
    button_xlsx.place(x=20,y=130)
    #-----------------------------------------------------------------
    button_dir = tk.Button(root, text='Select Directory',
                           font=("Helvetica", 10),
                           command=lambda: select_dir (button_xlsx))
    button_dir.place(x=20,y=70)
    #-----------------------------------------------------------------
    root.mainloop() #keep the window open while programme is running
    return

def user_guide():
    '''Open user guide documentation'''
    file = (Path(__file__).parent / "UserGuide" /"UserGuide.pdf").as_posix()
    webbrowser.open_new(file)
    return

def example():
    '''Copy a folder with examples to the user Desktop to try the programm'''
    #desktop = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop')
    desktop = os.path.normpath(os.path.expanduser("~/Desktop"))
    example_fold = (Path(__file__).parent / "Example_messi_nmr").as_posix()
    shutil.copytree(example_fold, os.path.join(desktop,"Example_messi_nmr"))

    examp_add = tk.Tk()
    examp_add.wm_title("MESSI")
    examp_add.geometry("400x100")

    label = tk.Label(examp_add, text='''An example folder had been created in your Desktop \n
Follow de User Guide instructions to procces its data.''')
    label.config(font = ("Times", "11"),justify='right')
    label.pack()
    tk.Button(examp_add, text='Ok',
              command=examp_add.destroy).pack(pady=10)

    #examp_add.mainloop()
    return

def select_dir(to_enable):
    '''Select directory. If it doesn't find .log and well rotulated
    elements doesnt enable next button.'''
    root.direc = filedialog.askdirectory(title='Select directory')
    os.chdir(root.direc)

    if not glob.glob('*_*.log'):
        tk.Label(root, text='G09 files not found. Try again').place(x=150,y=75)
        return

    tk.Label(root, text=root.direc).place(x=150,y=75)
    global isomer_list
    isomer_list, cant_comp = isomer_count() #this funtion globals "isomer_list"
    tk.Label(root, text=f'Isomeric Candidates: {isomer_list}').place(x=150,y=95)

    to_enable['state']='active'

    return

def select_xlsx(to_disable):
    '''Select xlsx file with experimental data and asignation labels.
    If it doesn't find "shifts" sheet, ask again and doesnt enable next button'''
    global xlsx
    xlsx = root.xlsx = filedialog.askopenfilename(title='Select Excel',
                                   filetypes=[('Excel files','*.xlsx'),
                                              ('All files','*')])

    xlsx_label = tk.Label(root)
    xlsx_label.place(x=150,y=135)

    final_label = tk.Label(root)
    final_label.place(x=150,y=155)

    if root.xlsx == '':
        xlsx_label.config( text='<- Select exp. data file                                     ')
        return

    elif 'shifts' not in pd.ExcelFile(root.xlsx).sheet_names:
        xlsx_label.config(text='"shifts" sheet wasn not found. Try again')
        root.xlsx = ''
        return

    else:
        root.after(2000, root.quit)

        final_label.config(text=f'Processing ...                      ')
        xlsx_label.config(text=root.xlsx)
        for button in to_disable:
            button['state']='disabled'

    return

def isomer_count():
    '''Determine the amount of isomeric candidates to be evaluated
    The files must be named by: isomerID_ * .log
    Funtion globals var "isomer_list" '''
    global isomer_list
    files= glob.glob('*.log') + glob.glob('*.out')
    isomer_list =[]
    for file in files:
        if file.split('_',1)[0] not in isomer_list:
            isomer_list.append(file.split('_',1)[0])
        else:
            continue
    isomer_list.sort() ##RG
    isomer_list.sort(key=lambda s: len(s)) #RG
    return isomer_list, len(isomer_list)

def GUI_end():
    '''Final window/widget to politely finish the programe'''
    byby = tk.Tk()
    byby.wm_title("MESSI")
    byby.geometry("150x100")

    tk.Label(byby, text='Process completed \nPress Exit.').place(x=20,y=20)
    tk.Button(byby, text='Exit',
              command=exit).place(x=60,y=60)

    byby.mainloop()
    return

def create_exe():
    '''Creates a direc acces executable file in the user desktop'''
    #desktop = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop')
    desktop = os.path.normpath(os.path.expanduser("~/Desktop"))
    dir_acc = os.path.join(desktop,'messi.py')

    with open (dir_acc, 'w') as file:
        file.write('# -*- coding: utf-8 -*-\n\n')
        file.write('import os, shutil\n\n')
        file.write('exe = shutil.which("messi")\n\n')
        file.write('os.system(exe)\n\n')


def change_directory(s):
    'Define work directory'
    print(f'\n{s:s}')
    root = Tk()
    root.withdraw()
    root.attributes('-topmost', True)
    di = filedialog.askdirectory()
    if di:
        os.chdir(di)
        print(f' -> Using the directory \n    {di:s}')
        return True
    else:
        print(' -> cancel.')
        return False

def data_sheet(open_file, cant_comp):
    'Allows you to read Excel with the label, interactively \ n'
    #print('\nSelect the excel file with the experimental data and labels.')
    #open_file = filedialog.askopenfilename()
    #if open_file=='':
    #    print(' -> cancel.')
    #    return False

    print(f' -> Using {open_file:s} as excel file with experimental data and labels.')

    df = pd.read_excel(open_file, sheet_name='shifts',engine='openpyxl')
    data = np.array(df[df['nuclei'].isna() == False])
    shifts = data
    d_exp_C = np.array([shifts[i][1:4] for i in range(shifts.shape[0]) if shifts[i][0] in 'cC'])
    d_exp_H = np.array([shifts[i][1:4] for i in range(shifts.shape[0]) if shifts[i][0] in 'hH'])
    if shifts.shape[1] < 7:
        wtl_C = np.array([shifts[i][4:6] for i in range(shifts.shape[0]) if shifts[i][0] in 'cC'])
        wtl_H = np.array([shifts[i][4:6] for i in range(shifts.shape[0]) if shifts[i][0] in 'hH'])
    else:
        for i in range(cant_comp):
            end_label = (cant_comp *3) + 4
            wtl_C = np.array([shifts[i][4:end_label] for i in range(shifts.shape[0]) if shifts[i][0] in 'cC'])
            wtl_H = np.array([shifts[i][4:end_label] for i in range(shifts.shape[0]) if shifts[i][0] in 'hH'])

    return d_exp_C, wtl_C, d_exp_H, wtl_H

def diasterotopics(tens_C, d_exp_C):
    '''If there are diasterotopic nuclei, the smallest value of the exp data
    will be matched with the smallest value of the calculations for each isomer.
    Returns the same variables corrected'''

    tens_C2 = tens_C.copy()
    d_exp_C2 = d_exp_C.copy()

    for i in range(d_exp_C.shape[0]-1):
        if d_exp_C2[i][2]!='nan':
            for j in range(i, d_exp_C.shape[0]-1):
                if d_exp_C2[i][2] == d_exp_C2[j+1][2]:
                    d_exp_C[i][1] = max(d_exp_C2[i][1], d_exp_C2[j+1][1])
                    d_exp_C[j+1][1] = min(d_exp_C2[i][1], d_exp_C2[j+1][1])
                    tens_C[i] = min(tens_C2[i], tens_C2[j+1])
                    tens_C[j+1]= max(tens_C2[i], tens_C2[j+1])
                else:
                    continue
        else:
            continue
    return tens_C, d_exp_C

def label_check(wtl, isom):
    '''Change de labeling if required'''
    try:
        if len(wtl[0]) < 4:
            return wtl
        else:
            start = (isom -1) *3
            end = start + 3
            return wtl[:, start:end]
    except:
        return wtl


def get_energy(file):
    'Extract SCF energies of every Gaussian 09 output'
    with open (file,'rt') as f:
        lines=f.readlines()
        for i, line in enumerate(lines):
            if "SCF Done:" in line:
                energy=float(line.split()[4])
    return energy

def relative_energies(energies):
    '''Receive a list with the energies of all the conformers and get the energy
    relative to the minimum'''
    energ = np.array(energies)
    energ *= 627.5095
    mas_estable = energ.min()
    e_relativas = energ - mas_estable
    return e_relativas

def get_information(file):
    '''Recive an output of gaussian 09 (.log o .out) and return 4 elements:
    The txt file in lines, SCF energy at RmPW1PW91 and RB3LYP levels,
    and a DataFrame of tensors,
    '''
    tensors=[]
    energy_PCM= 0
    energy_SMD= 0
    with open (file,'rt') as f:
        lines=f.readlines()
        for i, line in enumerate(lines):
            if "SCF Done:" and "E(RmPW1PW91)" in line:
                energy_PCM=float(line.split()[4])
            if "Isotropic = " in line:
                tensors.append(float(line.split()[4]))
            if "SCF Done:" and "E(RB3LYP)" in line:
                energy_SMD=float(line.split()[4])

    return lines, energy_PCM, energy_SMD, pd.DataFrame(tensors)

def Boltzman(tens_all_conf, e_relativas, Fb):
    '''Once all the tensors  have been extracted and corrected, this function
     performs the weighted averages according to the Boltzman probability and Boltzman factor.
     Its parameters require a DataFrame with the tensors of C and H
     a list of those relative energies and the boltzman factor (Fb).
    '''
    factor_bolt = Fb*0.25
    P_Boltz = (np.exp((-e_relativas*4.18*factor_bolt)/2.5))
    if Fb == 0:
        P_Boltz = np.empty(e_relativas.shape)
        for i, e in enumerate(np.nditer(e_relativas)):
            if e == 1000:
                P_Boltz[i] = 0
            else:
                P_Boltz[i] = 1
        contribucion = P_Boltz / P_Boltz.sum()
        tensores_ponderados = pd.DataFrame(tens_all_conf) * contribucion
        tensores_ponderados = tensores_ponderados.sum(axis=1)
    else:
        contribucion = P_Boltz / P_Boltz.sum()
        tensores_ponderados = pd.DataFrame(tens_all_conf) * contribucion
        tensores_ponderados = tensores_ponderados.sum(axis=1)

    return tensores_ponderados

def energy_converter(e_rel, filtro):
    DE_tot = max(e_rel)
    e_filtradas = []
    corte_superior = np.mean(e_rel) + 3**np.std(e_rel)
    corte_F2 = np.mean(e_rel) *2

    if filtro[1] == 1:
        DE_w = filtro[2]
    elif filtro[1] == 2:
        DE_w = 0.1*filtro[2]*corte_F2
    else:
        DE_w = np.mean(e_rel) - (3-0.5*filtro[2])*np.std(e_rel)


    for e in e_rel:
        if e < DE_w or e > corte_superior:
            e_filtradas.append(1000)
        else:
            e_filtradas.append(e)
    return np.array(e_filtradas)

def filters(tens_all_conf, energies_PCM, energies_SMD):
    '''Filter Parameters: DEw = remotion window DE = energy window
    Level of theory (A:PCM/mPW1PW91 B:SMD/B3LYP)
    Strategies of conformational removal (F = 1 (fixed), 2 (0.1nDE), 3(mu - 0.5n)*sigma)
    Remotion level n = (1-4)
    Boltzman Factor fb = 0-4
    1: [A,1,0,1], 2: [A, 1,0,2], 3: [A, 1,1,0], 4: [A,2,1,2],
    5: [A,3,3,2], 6: [A,3,3,3], 7: [B,1,0,0], 8: [B,1,0,1],
    9: [B,1,0,2], 10: [B,1,0,3], 11: [B,1,1,3], 12: [B,1,1,4],
    13: [B,2,1,3], 14: [B,2,1,4], 15: [B,3,1,2], 16: [B,3,3,2]
            '''
    A = 'PCM'
    B = 'SMD'
    #DP4_plus = [A,1,0,4]
    filters = [[A,1,0,4],
               [A,1,0,1], [A, 1,0,2],[A, 1,1,0],[A,2,1,2],
               [A,3,3,2], [A,3,3,3], [B,1,0,0], [B,1,0,1],
               [B,1,0,2], [B,1,0,3], [B,1,1,3], [B,1,1,4],
               [B,2,1,3], [B,2,1,4], [B,3,1,2], [B,3,3,2], [B,1,0,4]]

    e_rel_PCM = relative_energies(energies_PCM)
    e_rel_SMD = relative_energies(energies_SMD)
    #DE_PCM = max(e_rel_PCM)
    #DE_SMD = max(e_rel_SMD)
    filtered_tensors = []
    for i, filtro in enumerate(filters):
        if filtro[0] == A:
            e_rel = e_rel_PCM.copy()
        else:
            e_rel = e_rel_SMD.copy()
        #print(e_rel)
        e_filtradas = energy_converter(e_rel, filtro)
        #print(e_filtradas)
        tensores_ponderados =  Boltzman(tens_all_conf, e_filtradas, filtro[3])
        filtered_tensors.append(tensores_ponderados)
    return pd.DataFrame(filtered_tensors)

def tens_ordenados(tensors_indexed_C, wtl_C, wtl_H):
    ''' To operate you need the tensors indexed in Gaussian order and the
    labels of the compound in np.array format (this should have only 3 columns)
    The result is a list with the tensors ordered according to the inserted
    label'''
    tens_C = np.zeros(wtl_C.shape)
    tens_H = np.zeros(wtl_H.shape)
    for i in range (wtl_C.shape[0]):
        for j in range (3):
            if not isnan(wtl_C[i,j]):
                index = int(wtl_C[i,j])
                tens_C[i,j] = tensors_indexed_C[index-1]
            else:
                tens_C[i,j]=float('nan')
    for i in range (wtl_H.shape[0]):
        for j in range (3):
            if not isnan(wtl_H[i,j]):
                index = int(wtl_H[i,j])
                tens_H[i,j] = tensors_indexed_C[index-1]
            else:
                tens_H[i,j]=float('nan')
    tens_C = pd.DataFrame(tens_C)
    tens_C = tens_C.mean(axis=1)

    tens_H = pd.DataFrame(tens_H)
    tens_H = tens_H.mean(axis=1)
    return tens_C, tens_H

def d_calculation(tensors_C, tensors_H):
    '''It use the tensors for the calculation of the chemical shifts using
    TMS as a reference standard'''
    TMS_C = 196.8009 #sin agua 196.6095
    TMS_H = 31.5496 #sin agua 31.560
    Unscaled_shift_C = pd.DataFrame(TMS_C - tensors_C)
    Unscaled_shift_H = pd.DataFrame(TMS_H - tensors_H)
    return Unscaled_shift_C, Unscaled_shift_H

def escalado_CyH(Unscaled_shift, exp):
    '''Performs C and H scaling.
     You will also order interchangeable H or C but requiered that the
     experimental chemical shifts to exchange are labeled with the same character'''
    shifts = []
    exchange = []

    if exp.shape[0] != 0:
        for i in range(exp.shape[0]):
            shifts.append(exp[i][1])
            exchange.append(exp[i][2])
    else:
        return pd.DataFrame(exp), pd.DataFrame(exp)

    UnsC_d = Unscaled_shift.copy()
    indices_intercambiables = [i for i in range(len(exchange)) if exchange[i] ==1]
    for e in indices_intercambiables:
        UnsC_d[e] = max(Unscaled_shift[e], Unscaled_shift[e+1])
        UnsC_d[e+1] = min(Unscaled_shift[e], Unscaled_shift[e+1])


    shifts = np.array(shifts).reshape(-1,1)
    UnsC_d = np.array(UnsC_d).reshape(-1,1)
    regresion = linear_model.LinearRegression()

    if len(shifts) == 1:
        Scaled_shift = shifts
    else:
        regresion.fit(shifts, UnsC_d)

        m = regresion.coef_
        b = regresion.intercept_

        Scaled_shift = (UnsC_d - b) / m
    Scaled_shift = pd.DataFrame(Scaled_shift)

    shifts = pd.DataFrame(shifts)
    return Scaled_shift, shifts

def error_calculator(calc, exp):
    return np.array(calc - exp)

def hibridization_classifier(err, d_exp):
    UscE_sp2 = pd.DataFrame()
    UscE_sp3 = pd.DataFrame()
    err = np.array(err)
    for i in range(d_exp.shape[0]):
        if d_exp[i][0] == 1:
            UscE_sp2 = pd.concat([UscE_sp2, pd.DataFrame(err[i])], axis = 1)
        else:
            UscE_sp3 =pd.concat([UscE_sp3, pd.DataFrame(err[i])], axis = 1)
    return UscE_sp2.transpose(), UscE_sp3.transpose()

def probability(error, mu, sigma, nu):
        t_dist = stats.t(nu)
        prob = 1 - t_dist.cdf(np.abs(error - mu) / sigma)
        return prob

def DP4_plus_calculator(errors_list, cant_comp):
    '''Function that calculate the DP4 probability'''
    #errors_list = [scEC, scEH, UscEC_sp2, UscEH_sp2, UscEC_sp3, UscEH_sp3]
    '''dist_parameters = {0: (0, 1.5569, 6.2274), 1: (0, 0.1044, 3.8935),
                       2:(-0.9201, 1.7484, 5.3643), 3: (0.3475, 0.1176, 4.911),
                       4: (2.9085, 1.6, 6.2693), 5:(-0.018, 0.1118, 3.6511)}'''
    dist_parameters = {0: (0, 1.5569, 6), 1: (0, 0.1044, 3),
                       2:(-0.9201, 1.7484, 5), 3: (0.3475, 0.1176, 4),
                       4: (2.9085, 1.6, 6), 5:(-0.018, 0.1118, 3)}

    #DP4_types_names = ['sDP4_C', 'sDP4_H', 'sDP4_Full', 'UsDP4_C', 'UsDP4_H', 'UsDP4_Full',
                 #'DP4_C', 'DP4_H', 'DP4_Full']
    probabilities = []
    for i, error_type in enumerate(errors_list):
        mu, sigma, nu = dist_parameters[i]
        if len(error_type) > 0:
            p_e = probability(error_type, mu, sigma, nu)
        else:
            p_e = np.ones((1, cant_comp))
        p_e = np.prod(p_e,0)
        probabilities.append(p_e)

    #errors_list = [scEC, scEH, UscEC_sp2, UscEH_sp2, UscEC_sp3, UscEH_sp3]
    sDP4_H = np.array(probabilities[1])
    sDP4_C = np.array(probabilities[0])
    sDP4_Full = np.array([probabilities[0], probabilities[1]])
    UnsDP4_H = np.array([probabilities[3], probabilities[5]])
    UnsDP4_C = np.array([probabilities[2], probabilities[4]])
    UnsDP4_Full = np.array([probabilities[3], probabilities[5], probabilities[2], probabilities[4]])
    DP4_plus_H = np.array([probabilities[1], probabilities[3], probabilities[5]])
    DP4_plus_C = np.array([probabilities[0], probabilities[2], probabilities[4]])
    DP4_plus_Full = np.array(probabilities)

    DP4_types = [sDP4_H, sDP4_C, sDP4_Full, UnsDP4_H, UnsDP4_C, UnsDP4_Full,
                 DP4_plus_H, DP4_plus_C, DP4_plus_Full]
    DP4_results = pd.DataFrame()

    for i, prob in enumerate(DP4_types):
        if np.prod(prob) == 1:
            DP4 = pd.DataFrame(np.array(['-' for _ in range(cant_comp)]))
        else:
            if i <2:
                DP4 = pd.DataFrame(100*prob/sum(prob))
            else:
                DP4 = pd.DataFrame(100*np.prod(prob,0)/sum(np.prod(prob,0)))

        DP4_results = pd.concat([DP4_results, DP4], axis=1)

    return DP4_results

def DP4_filter_calculation(i, cant_comp, tens_matrix_all_isomers, wtl_C_all, wtl_H_all, d_exp_C, d_exp_H):
    UnS_shifts_C = pd.DataFrame()
    UnS_shifts_H = pd.DataFrame()
    isomers_shifts_C = pd.DataFrame()
    isomers_shifts_H = pd.DataFrame()
    isom_sin_conf = []
    for j in range(cant_comp):
        wtl_C = label_check(wtl_C_all, j+1) #labels verifier
        wtl_H = label_check(wtl_H_all, j+1)
        isom_tens = tens_matrix_all_isomers[j].iloc[i]
        tens_C, tens_H = tens_ordenados(isom_tens, wtl_C, wtl_H)
        '''Diastereotopic order'''
        tens_C, d_exp_C = diasterotopics(tens_C, d_exp_C)
        tens_H, d_exp_H = diasterotopics(tens_H, d_exp_H)
        '''Chemical shifts calculation and Scaling procedure '''
        if tens_C.sum() == 0 and tens_H.sum() == 0:
            Unscaled_shift_C, Unscaled_shift_H = pd.DataFrame(196.8009-tens_C), pd.DataFrame(31.5496-tens_H)
            Scaled_shift_C, Scaled_shift_H = pd.DataFrame(196.8009-tens_C), pd.DataFrame(31.5496-tens_H)
            isom_sin_conf.append((j, i))
        else:
            Unscaled_shift_C, Unscaled_shift_H = d_calculation(tens_C, tens_H)
            Scaled_shift_C, exp_C = escalado_CyH(Unscaled_shift_C, d_exp_C)
            Scaled_shift_H, exp_H = escalado_CyH(Unscaled_shift_H, d_exp_H)

        '''Once the corrected tensors and scaled chemical shifts of an isomer
        are obtained, they are added to the DataFrame of all the isomers that
        will be correlated with the experimental data '''

        UnS_shifts_C = pd.concat([UnS_shifts_C, Unscaled_shift_C ],axis=1)
        UnS_shifts_H = pd.concat([UnS_shifts_H, Unscaled_shift_H ],axis=1)

        isomers_shifts_C = pd.concat([isomers_shifts_C, Scaled_shift_C ],axis=1)
        isomers_shifts_H = pd.concat([isomers_shifts_H, Scaled_shift_H ],axis=1)

    '''Unscaled Error calculation'''
    UscEC = pd.DataFrame(error_calculator(UnS_shifts_C, exp_C))
    UscEH = pd.DataFrame(error_calculator(UnS_shifts_H, exp_H))

    '''Sp2 and Sp3 classifier'''
    UscEC_sp2, UscEC_sp3 = hibridization_classifier(UscEC, d_exp_C)
    UscEH_sp2, UscEH_sp3 = hibridization_classifier(UscEH, d_exp_H)

    '''Calculation of the scaled H and C errors'''
    scEC = error_calculator(isomers_shifts_C, exp_C)
    scEH = error_calculator(isomers_shifts_H, exp_H)

    errors_list = [scEC, scEH, UscEC_sp2, UscEH_sp2, UscEC_sp3, UscEH_sp3]

    '''Calculation of the partial or complete DP4_plus multiensambles'''
    DP4_results = DP4_plus_calculator(errors_list, cant_comp)
    return DP4_results, pd.DataFrame(isom_sin_conf)

def isomers_names(isomer_list, cant_comp):
    isomers = []
    for isom in isomer_list:
        isomer = 'Isomer '
        for caracter in isom:
            if caracter != '_':
                isomer += caracter
            else:
                break
        isomers.append(isomer)
    return isomers

def modify_report(report, problems):
    b = len(problems)
    #ABC_DARIO=['A','B','C',D,E,F,G,H,I,J,K,L,M,N,O,P,Q,R,S,T,U,V,W,X,Y,Z]
    ABC_DARIO = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
    wb = load_workbook(report)
    ws = wb['Results']
    ws.column_dimensions['A'].width = 20
    ws['A2'].fill = PatternFill(start_color = '00CED1', end_color = '00CED1', fill_type = "solid")
    for i in range(3, 19):
        ws[f'A{i}'].fill = PatternFill(start_color = 'AFEEEE', end_color = 'AFEEEE', fill_type = "solid")
    ws['A19'].fill = PatternFill(start_color = 'FFE4E1', end_color = 'FFE4E1', fill_type = "solid")
    ws['A20'].fill = PatternFill(start_color = 'F5DEB3', end_color = 'F5DEB3', fill_type = "solid")
    for i in range(b):
        row = problems.iloc[i,1]+2
        col = ABC_DARIO[problems.iloc[i,0] + 1]
        ws[f'{col}{row}'].fill = PatternFill(start_color = 'B0C4DE', end_color = 'B0C4DE', fill_type = "solid")

    wb.save('MESSI_Results.xlsx')
    return

def main():
    print(f'In the directory {(Path(__file__).parent / "examples").as_posix()} you can find examples.')

    GUI() #it globals xlsx (the calc sheet with the input data)

    #if change_directory('Select the input data directory.')==False:
    #   return

    #ds = data_sheet(cant_comp)
    ds = data_sheet(xlsx, len(isomer_list))
    cant_comp = len(isomer_list)

    # Reading data
    #cant_comp, isomer_list = isomer_count()

    if ds==False:
        return
    d_exp_C, wtl_C_all, d_exp_H, wtl_H_all  = ds

    print("\nProcessing...")
    isomers_tensors = pd.DataFrame()

    tens_matrix_all_isomers = []

    for n_isom, isom in enumerate(isomer_list):

        conformers = glob.glob(f'{isom}*.log') +  glob.glob(f'{isom}*.out')
        conformers.sort() ##RG
        conformers.sort(key=lambda s: len(s)) #RG

        wtl_C = label_check(wtl_C_all, n_isom+1) #labels verifier
        wtl_H = label_check(wtl_H_all, n_isom+1)

        tens_all_conf = pd.DataFrame()
        energies_PCM=[]
        energies_SMD=[]
        for n,conf in enumerate(conformers):
            '''Both energies and the tensors are extracted from each conformer.
            Both energies could be in one or separeted output files'''
            lines, energy_PCM, energy_SMD, tensors = get_information(conf)
            if energy_PCM != 0:
                energies_PCM.append(energy_PCM)
            if energy_SMD != 0:
                energies_SMD.append(energy_SMD)
            tens_all_conf = pd.concat([tens_all_conf,tensors],axis=1)

        '''Obtaining a vector with an applied filter, Boltzmann-weighted corrected tensors'''
        matrix_tensors = filters(tens_all_conf, energies_PCM, energies_SMD)
        tens_matrix_all_isomers.append(matrix_tensors)

    DP4_all_filters = pd.DataFrame()
    DP4_plus_all_filters = pd.DataFrame()
    problems=pd.DataFrame()
    for i in range(len(matrix_tensors)):
        DP4_results, isom_sin_conf = DP4_filter_calculation(i, cant_comp, tens_matrix_all_isomers, wtl_C_all, wtl_H_all, d_exp_C, d_exp_H)
        DP4_plus = pd.DataFrame(np.array(DP4_results.T)[8])
        problems = pd.concat([problems, isom_sin_conf], axis=0)
        if i ==0:
            DP4_plus_original=  DP4_plus
        elif i == 17:
            DP4_plus_SMD=  DP4_plus
        else:
            DP4_plus_all_filters = pd.concat([ DP4_plus_all_filters, DP4_plus], axis=1)
            DP4_all_filters= pd.concat([DP4_all_filters, DP4_results], axis=0)


    "Data procesing to print data to an excel file"
    isomers = isomers_names(isomer_list, cant_comp)
    messi = DP4_plus_all_filters.mean(axis=1)

    filters_type = pd.DataFrame(['MESSI', '1: [A,1,0,1]', '2: [A,1,0,2]', '3: [A,1,1,0]', '4: [A,2,1,2]','5: [A,3,3,2]', '6: [A,3,3,3]', '7: [B,1,0,0]', '8: [B,1,0,1]', '9: [B,1,0,2]', '10: [B,1,0,3]', '11: [B,1,1,3]', '12: [B,1,1,4]', '13: [B,2,1,3]', '14: [B,2,1,4]', '15: [B,3,1,2]', '16: [B,3,3,2]', 'DP4+_PCM', 'DP4+_SMD'])
    a = pd.concat([messi, DP4_plus_all_filters, DP4_plus_original, DP4_plus_SMD], axis=1).T
    DP4_results_all = a.reset_index(drop=True)
    DP4_results_all = pd.concat([filters_type, DP4_results_all], axis=1)
    DP4_results_all.columns= ['Filter']+isomers


    print('\nDone!')
    perm=False

    while not perm:
        perm=True
        try:

            directorio = os.getcwd()
            workbook = os.path.join(directorio,'MESSI_Results.xlsx')
            print(f'\n -> Writting output file MESSI_Results.xlsx in {directorio:s}.')
            with open(workbook, 'w') as f:
                f.write('output file')
        except:
            print("   -> Can't write. Please choose another directory for output file.")
            perm=False
            if change_directory('Select the output directory.')==False:
                print('Exit.')
                return

    label_columns = [n+1 for n in range(matrix_tensors.shape[1])]
    with pd.ExcelWriter(workbook) as writer:
        DP4_results_all.to_excel(writer,sheet_name='Results', index=False,float_format="%.2f")
        for i, isom in enumerate(isomers):
            tens_matrix_all_isomers[i].columns = label_columns
            tens_matrix_all_isomers[i].to_excel(writer,sheet_name=f'Ten_{isom}',index=False, float_format="%.4f")


    modify_report('MESSI_Results.xlsx', problems)

    GUI_end()

if __name__=='__main__' :
    main()
